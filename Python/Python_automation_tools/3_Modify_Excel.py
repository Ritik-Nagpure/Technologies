import openpyxl
import os

filedir = str(os.getcwd()) + "\\00_Excels"
os.chdir(filedir)
print(os.getcwd())

# open workbook
wb = openpyxl.load_workbook('example.xlsx')
wb.create_sheet('newSheet')
ws = wb['Sheet2']
ws.title = "edited sheet"
ws = wb["Sheet1"]
ws["D1"] = "New Data"
ws["D3"] = "Cat"
ws["D5"] = "Dog"
wb.save('newexample.xlsx')
