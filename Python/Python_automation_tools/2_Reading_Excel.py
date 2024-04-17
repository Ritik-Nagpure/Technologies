import openpyxl
import os

# documentation ==> https://openpyxl.readthedocs.io/en/stable/tutorial.html
filedir = str(os.getcwd()) + "\\00_Excels"
os.chdir(filedir)
print(os.getcwd())

# open workbook
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
print(workbook)
print(workbook.sheetnames)
workbook.create_sheet('newSheet')
print(workbook.sheetnames)

activesheet = workbook['Sheet1']
print(type(activesheet))
print(activesheet)
activesheet.title = "edited sheet"
print(workbook.sheetnames)

currcell = activesheet["A4"]
print(currcell, type(currcell.value), currcell.value)

thiscell = activesheet.cell(row=4, column=2)
print(thiscell, type(thiscell.value), thiscell.value)
